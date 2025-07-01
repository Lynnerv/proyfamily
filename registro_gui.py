import tkinter as tk
from tkinter import messagebox, ttk
from openpyxl import Workbook, load_workbook
from datetime import datetime
import os

# ==================== 📁 CREAR ARCHIVO EXCEL SI NO EXISTE ====================
archivo_excel = "cumpleaños.xlsx"
if not os.path.exists(archivo_excel):
    wb = Workbook()
    ws = wb.active
    ws.title = "Cumpleaños"
    ws.append(["Nombre", "Apellido Paterno", "Apellido Materno", "Código País", "Número", "Fecha de Nacimiento"])
    wb.save(archivo_excel)

# ==================== 💾 FUNCIÓN PARA GUARDAR DATOS ====================
def guardar_datos():
    nombre = entry_nombre.get().strip()
    apellido_pat = entry_apellido_pat.get().strip()
    apellido_mat = entry_apellido_mat.get().strip()
    codigo = combo_codigo.get().split()[0].strip()
    numero = entry_numero.get().strip()
    dia = combo_dia.get().zfill(2).strip()
    mes_input = combo_mes.get().strip()
    anio = combo_anio.get().strip()

    if not (nombre and apellido_pat and apellido_mat and codigo and numero and dia and mes_input and anio):
        messagebox.showerror("❌ Error", "Todos los campos son obligatorios.")
        return

    meses = {
        "enero": "01", "febrero": "02", "marzo": "03", "abril": "04",
        "mayo": "05", "junio": "06", "julio": "07", "agosto": "08",
        "septiembre": "09", "octubre": "10", "noviembre": "11", "diciembre": "12"
    }

    mes_num = ""
    if mes_input.isdigit() and 1 <= int(mes_input) <= 12:
        mes_num = str(mes_input).zfill(2)
    else:
        mes_input_normalizado = mes_input.lower()
        if mes_input_normalizado in meses:
            mes_num = meses[mes_input_normalizado]
        else:
            messagebox.showerror("❌ Error", "Mes inválido. Usa nombre o número del mes válido.")
            return

    fecha_nac = f"{anio}-{mes_num}-{dia}"

    try:
        datetime.strptime(fecha_nac, "%Y-%m-%d")
    except ValueError:
        messagebox.showerror("❌ Error", "Fecha de nacimiento inválida.")
        return

    wb = load_workbook(archivo_excel)
    ws = wb.active
    ws.append([nombre, apellido_pat, apellido_mat, codigo, numero, fecha_nac])
    wb.save(archivo_excel)

    messagebox.showinfo("✅ Éxito", f"Datos de {nombre} guardados.")
    entry_nombre.delete(0, tk.END)
    entry_apellido_pat.delete(0, tk.END)
    entry_apellido_mat.delete(0, tk.END)
    entry_numero.delete(0, tk.END)
    combo_codigo.set('')
    combo_dia.set('')
    combo_mes.set('')
    combo_anio.set('')

# ==================== 🎨 ESTILO ELEGANTE CON FUENTE GRANDE ====================
ventana = tk.Tk()
ventana.title("🎂 Registro de Cumpleaños")
ventana.geometry("440x580")  # Altura aumentada para que se vea el botón

# Fuente global
fuente_general = ("Segoe UI", 12)
fuente_boton = ("Segoe UI", 12, "bold")

# Colores
fondo_color = "#1f1f1f"      # Grafito oscuro
campo_color = "#2c2c2c"
texto_color = "#f5f5f5"
boton_color = "#d4af37"      # Dorado elegante
boton_hover = "#f0c846"

ventana.configure(bg=fondo_color)

# Estilos
style = ttk.Style()
style.theme_use("clam")

style.configure("TLabel", font=fuente_general, background=fondo_color, foreground=texto_color)
style.configure("TEntry", font=fuente_general, fieldbackground=campo_color, foreground=texto_color)
style.configure("TCombobox", font=fuente_general, fieldbackground=campo_color, background=campo_color, foreground=texto_color)
style.configure("TButton", font=fuente_boton, background=boton_color, foreground="black", borderwidth=0)
style.map("TButton",
          background=[("active", boton_hover)],
          foreground=[("active", "black")])

def crear_label(texto):
    return ttk.Label(ventana, text=texto)

def espaciado(px=5):
    tk.Label(ventana, text="", bg=fondo_color).pack(pady=px)

# ==================== 📋 CAMPOS DEL FORMULARIO ====================
crear_label("🧑 Nombre").pack()
entry_nombre = ttk.Entry(ventana, width=30, font=fuente_general)
entry_nombre.pack()

espaciado()

crear_label("👨 Apellido Paterno").pack()
entry_apellido_pat = ttk.Entry(ventana, width=30, font=fuente_general)
entry_apellido_pat.pack()

espaciado()

crear_label("👩 Apellido Materno").pack()
entry_apellido_mat = ttk.Entry(ventana, width=30, font=fuente_general)
entry_apellido_mat.pack()

espaciado()

crear_label("🌐 Código de país").pack()
combo_codigo = ttk.Combobox(ventana, values=[
    "+51 (Perú)", "+52 (México)", "+54 (Argentina)", "+57 (Colombia)", "+1 (EEUU/Canadá)"
], width=27, font=fuente_general)
combo_codigo.pack()

espaciado()

crear_label("📞 Número").pack()
entry_numero = ttk.Entry(ventana, width=30, font=fuente_general)
entry_numero.pack()

espaciado()

crear_label("🎉 Fecha de nacimiento").pack()
frame_fecha = tk.Frame(ventana, bg=fondo_color)
frame_fecha.pack()

combo_dia = ttk.Combobox(frame_fecha, values=[str(i) for i in range(1, 32)], width=5, font=fuente_general)
combo_dia.pack(side=tk.LEFT, padx=4)

combo_mes = ttk.Combobox(frame_fecha, values=[
    "Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio",
    "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre",
    "1", "2", "3", "4", "5", "6", "7", "8", "9", "10", "11", "12"
], width=10, font=fuente_general)
combo_mes.pack(side=tk.LEFT, padx=4)

combo_anio = ttk.Combobox(frame_fecha, values=[str(i) for i in range(1920, datetime.now().year + 1)], width=8, font=fuente_general)
combo_anio.pack(side=tk.LEFT, padx=4)

espaciado(10)

# ==================== 💾 BOTÓN GUARDAR ====================
ttk.Button(ventana, text="💾 Guardar registro", command=guardar_datos).pack(pady=15)

# ==================== 🚀 INICIAR INTERFAZ ====================
ventana.mainloop()
